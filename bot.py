import os
import re
import asyncio
import logging
import shutil
import json
from datetime import time as dtime, datetime
from zoneinfo import ZoneInfo
from telegram import Update, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    CallbackQueryHandler, filters, ContextTypes,
)
from dotenv import load_dotenv
import excel_handler

logging.basicConfig(format="%(asctime)s %(levelname)s %(message)s", level=logging.INFO)
logging.getLogger("httpx").setLevel(logging.WARNING)  # suppress HTTP logs that expose bot token

_WATCHED = [
    "TELEGRAM_BOT_TOKEN", "EXCEL_PATH",
    "GOOGLE_CLIENT_ID", "GOOGLE_CLIENT_SECRET",
    "GOOGLE_REFRESH_TOKEN", "GOOGLE_DRIVE_FOLDER_ID",
]

# Log what Railway actually injected into the process before we touch dotenv
logging.info("ENV before dotenv: %s",
             {k: ("SET" if k in os.environ else "MISSING") for k in _WATCHED})

# load_dotenv never overrides existing os.environ values (override=False default),
# so Railway-injected vars always win; /data/.env fills in only what's missing.
load_dotenv("/data/.env")
load_dotenv()

logging.info("ENV after  dotenv: %s",
             {k: ("SET" if os.getenv(k) else "MISSING") for k in _WATCHED})

TOKEN = os.getenv("TELEGRAM_BOT_TOKEN")
PAGE_SIZE = 8
CHAT_ID_FILE = os.path.join(os.path.dirname(__file__), "chat_id.txt")
ISRAEL_TZ = ZoneInfo("Asia/Jerusalem")

PAY_LABELS = {"cash": "💵 מזומן", "credit": "💳 אשראי"}


def save_chat_id(chat_id: int):
    with open(CHAT_ID_FILE, "w") as f:
        f.write(str(chat_id))


def load_chat_id():
    if os.path.exists(CHAT_ID_FILE):
        with open(CHAT_ID_FILE) as f:
            return int(f.read().strip())
    return None


def _chat_ids_path() -> str:
    ep = os.getenv("EXCEL_PATH", "")
    base = os.path.dirname(ep) if ep else os.path.dirname(os.path.abspath(__file__))
    return os.path.join(base, "chat_ids.json")


def _load_chat_ids_data() -> dict:
    path = _chat_ids_path()
    if os.path.exists(path):
        try:
            with open(path) as f:
                return json.load(f)
        except Exception:
            pass
    return {"ids": [], "names": {}}


def register_chat_id(chat_id: int, name: str = ""):
    data = _load_chat_ids_data()
    if chat_id not in data["ids"]:
        data["ids"].append(chat_id)
    if name:
        data.setdefault("names", {})[str(chat_id)] = name
    try:
        with open(_chat_ids_path(), "w") as f:
            json.dump(data, f)
    except Exception as e:
        logging.warning("Failed to save chat_ids: %s", e)


def load_all_chat_ids() -> list:
    return _load_chat_ids_data().get("ids", [])


def get_user_name(chat_id: int) -> str:
    data = _load_chat_ids_data()
    return data.get("names", {}).get(str(chat_id), "")


def fmt(n):
    return f"{n:,.0f}"


def format_status(expenses, income, month_name):
    total_income   = sum(c["actual"] for c in income)
    total_expenses = sum(c["actual"] for c in expenses)
    diff = total_income - total_expenses
    diff_icon = "📈" if diff >= 0 else "📉"
    diff_sign = "+" if diff >= 0 else ""

    lines = [
        f"\U0001f4ca *דוח {month_name}*\n",
        f"💰 הכנסות (בעל + אישה): *{fmt(total_income)} ₪*",
        f"💸 הוצאות: *{fmt(total_expenses)} ₪*",
        f"{diff_icon} הפרש: *{diff_sign}{fmt(diff)} ₪*",
    ]

    if expenses:
        budgeted  = [c for c in expenses if c.get("budget", 0) > 0]
        if budgeted:
            total_bud = sum(c["budget"] for c in budgeted)
            total_act = sum(c["actual"] for c in budgeted)
            overruns  = [c for c in budgeted if c["actual"] > c["budget"]]
            near      = [c for c in budgeted if 0 < c["budget"] - c["actual"] < c["budget"] * 0.15]
            ok_count  = len(budgeted) - len(overruns) - len(near)
            pct_used  = int(total_act / total_bud * 100) if total_bud > 0 else 0
            lines.append("")
            lines.append("*📌 סיכום תקציב:*")
            lines.append(f"   ניצול: *{fmt(total_act)} / {fmt(total_bud)} ₪*  ({pct_used}%)")
            status_parts = []
            if overruns:  status_parts.append(f"🔴 {len(overruns)} חריגות")
            if near:      status_parts.append(f"🟡 {len(near)} קרוב לגבול")
            if ok_count:  status_parts.append(f"🟢 {ok_count} בתקציב")
            if status_parts:
                lines.append("   " + "  |  ".join(status_parts))

        lines.append("")
        lines.append("*הוצאות לפי קטגוריה:*")
        for c in sorted(expenses, key=lambda x: -x["actual"]):
            budget = c.get("budget", 0)
            actual = c["actual"]
            name   = c["name"]
            if budget > 0:
                remaining = budget - actual
                pct = actual / budget
                if remaining < 0:
                    lines.append(f"  🔴 {name}: {fmt(actual)} / {fmt(budget)} ₪  ⚠️ +{fmt(abs(remaining))}")
                elif pct >= 0.85:
                    lines.append(f"  🟡 {name}: {fmt(actual)} / {fmt(budget)} ₪  ({fmt(remaining)} נותר)")
                else:
                    lines.append(f"  🟢 {name}: {fmt(actual)} / {fmt(budget)} ₪  ({fmt(remaining)} נותר)")
            else:
                lines.append(f"  ▪️ {name}: {fmt(actual)} ₪")

    return "\n".join(lines)


def format_expense_result(name, amount, result, pay_type=None, month_name=None):
    pay_label = PAY_LABELS.get(pay_type, "") if pay_type else ""
    month_str = f" ({month_name.strip()})" if month_name else ""
    budget    = result.get("budget", 0)
    actual    = result.get("actual", 0)
    remaining = result.get("remaining", 0)

    if budget > 0:
        pct = actual / budget
        if remaining < 0:
            status_icon = "🔴"
        elif pct >= 0.85:
            status_icon = "🟡"
        else:
            status_icon = "🟢"
    else:
        status_icon = "✅"

    line1 = f"{status_icon} נוספו *{fmt(amount)} ₪* לקטגוריית *{name}*{month_str}"
    if pay_label:
        line1 += f" — {pay_label}"
    lines = [line1, ""]

    if budget > 0:
        pct_display = int(pct * 100)
        lines.append(f"💰 תקציב: {fmt(budget)} ₪  |  📊 בוצע: {fmt(actual)} ₪  ({pct_display}%)")
        if remaining < 0:
            lines.append(f"⚠️ *חריגה של {fmt(abs(remaining))} ₪ מהתקציב!*")
        elif pct >= 0.85:
            lines.append(f"🟡 נותרו: *{fmt(remaining)} ₪* — מתקרב לתקציב")
        else:
            lines.append(f"🟢 נותרו: *{fmt(remaining)} ₪*")
    else:
        lines.append(f"📊 סה\"כ בקטגוריה: *{fmt(actual)} ₪*")

    return "\n".join(lines)


def make_pay_keyboard(row, amount):
    return InlineKeyboardMarkup([
        [
            InlineKeyboardButton("💳 אשראי", callback_data=f"pay:credit:{row}:{amount}"),
            InlineKeyboardButton("💵 מזומן",  callback_data=f"pay:cash:{row}:{amount}"),
        ],
        [InlineKeyboardButton("❌ ביטול", callback_data="cancel")],
    ])


def make_budget_cat_keyboard(cats, page=0):
    sorted_cats = sorted(cats, key=lambda x: x["name"])
    chunk = sorted_cats[page * PAGE_SIZE: (page + 1) * PAGE_SIZE]
    rows = []
    for c in chunk:
        budget = c.get("budget", 0)
        bstr   = f"{fmt(budget)}₪" if budget > 0 else "ללא יעד"
        label  = f"{c['name'][:18]} [{bstr}]"
        rows.append([InlineKeyboardButton(label[:32], callback_data=f"setbud:{c['row']}")])
    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton("◀️ הקודם", callback_data=f"budpage:{page - 1}"))
    if (page + 1) * PAGE_SIZE < len(sorted_cats):
        nav.append(InlineKeyboardButton("הבא ▶️",   callback_data=f"budpage:{page + 1}"))
    if nav:
        rows.append(nav)
    rows.append([InlineKeyboardButton("❌ ביטול", callback_data="cancel")])
    return InlineKeyboardMarkup(rows)


def make_budget_month_keyboard(row, amount):
    current_month = datetime.now().month
    months = excel_handler.MONTH_SHEETS
    rows = []
    row_btns = []
    for num in range(1, 13):
        label = months[num].strip()
        if num == current_month:
            label = f"• {label} •"
        row_btns.append(InlineKeyboardButton(
            label, callback_data=f"budmonth:{num}:{row}:{amount}"
        ))
        if len(row_btns) == 3:
            rows.append(row_btns)
            row_btns = []
    if row_btns:
        rows.append(row_btns)
    rows.append([InlineKeyboardButton("❌ ביטול", callback_data="cancel")])
    return InlineKeyboardMarkup(rows)


def make_month_keyboard(pay_type, row, amount):
    current_month = datetime.now().month
    months = excel_handler.MONTH_SHEETS
    rows = []
    row_btns = []
    for num in range(1, 13):
        label = months[num].strip()
        if num == current_month:
            label = f"• {label} •"
        row_btns.append(InlineKeyboardButton(
            label, callback_data=f"month:{num}:{pay_type}:{row}:{amount}"
        ))
        if len(row_btns) == 3:
            rows.append(row_btns)
            row_btns = []
    if row_btns:
        rows.append(row_btns)
    rows.append([InlineKeyboardButton("❌ ביטול", callback_data="cancel")])
    return InlineKeyboardMarkup(rows)


async def send_daily_report(context: ContextTypes.DEFAULT_TYPE):
    chat_ids = load_all_chat_ids()
    if not chat_ids:
        fallback = load_chat_id()
        if fallback:
            chat_ids = [fallback]
    if not chat_ids:
        return
    try:
        expenses, income, month = excel_handler.get_status()
    except Exception as e:
        logging.error("daily report failed: %s", e)
        return
    if not expenses and not income:
        return
    text = format_status(expenses, income, month)
    for cid in chat_ids:
        try:
            await context.bot.send_message(chat_id=cid, text=text, parse_mode="Markdown")
        except Exception as e:
            logging.warning("Daily report to %s failed: %s", cid, e)


def make_cat_keyboard(cats, amount, page=0):
    # Sort: categories with existing expenses first, then alphabetically
    sorted_cats = sorted(cats, key=lambda x: (-x["actual"], x["name"]))
    chunk = sorted_cats[page * PAGE_SIZE: (page + 1) * PAGE_SIZE]

    rows = []
    pair = []
    for c in chunk:
        label = c["name"][:22]
        pair.append(InlineKeyboardButton(label, callback_data=f"cat:{c['row']}:{amount}"))
        if len(pair) == 2:
            rows.append(pair); pair = []
    if pair:
        rows.append(pair)

    nav = []
    if page > 0:
        nav.append(InlineKeyboardButton("◀️ הקודם", callback_data=f"page:{page-1}:{amount}"))
    if (page + 1) * PAGE_SIZE < len(sorted_cats):
        nav.append(InlineKeyboardButton("הבא ▶️",   callback_data=f"page:{page+1}:{amount}"))
    if nav:
        rows.append(nav)

    return InlineKeyboardMarkup(rows)


async def cmd_start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    chat_id = update.effective_chat.id
    first_name = (update.effective_user.first_name or "") if update.effective_user else ""
    save_chat_id(chat_id)
    register_chat_id(chat_id, first_name)
    await update.message.reply_text(
        "שלום! אני בוט ניהול התקציב המשפחתי \U0001f4b0\n\n"
        "פקודות:\n"
        "• /status - דוח מצב חודשי\n"
        "• /dashboard - ויזואליזציה + ניתוח חריגות\n"
        "• /budget - עדכון יעדי תקציב לפי קטגוריה\n"
        "• /export - הורד קובץ האקסל\n\n"
        "הוספת הוצאה:\n"
        "• שלח `קטגוריה סכום` (לדוגמה: `דלק 150`)\n"
        "• או רק סכום ואני אציג רשימת קטגוריות\n\n"
        "✅ נרשמת לדוח יומי אוטומטי בשעה 08:00"
    )


async def cmd_dashboard(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        expenses, income, month = excel_handler.get_status()
    except Exception as e:
        await update.message.reply_text(f"⚠️ שגיאה: {e}")
        return

    active = [c for c in expenses if c["actual"] > 0]
    if not active:
        await update.message.reply_text("אין נתוני הוצאות לחודש הנוכחי")
        return

    active.sort(key=lambda x: -x["actual"])
    total_exp = sum(c["actual"] for c in active)
    total_inc = sum(c["actual"] for c in income)
    diff = total_inc - total_exp
    pct = (total_exp / total_inc * 100) if total_inc > 0 else 0
    diff_icon = "📈" if diff >= 0 else "📉"
    diff_sign = "+" if diff >= 0 else ""
    max_val = active[0]["actual"] if active else 1
    BAR_W = 9

    lines = [
        f"📊 *דוח {month}*\n",
        f"💰 הכנסות:  *{fmt(total_inc)} ₪*",
        f"💸 הוצאות:  *{fmt(total_exp)} ₪*  ({pct:.0f}%)",
        f"{diff_icon} נותר:     *{diff_sign}{fmt(diff)} ₪*",
        "",
        "━━━━━━━━━━━━━━━━━━━━━━",
    ]

    RLM = "‏"  # Right-to-Left Mark — forces RTL paragraph direction

    overruns = []
    for c in active[:15]:
        actual = c["actual"]
        budget = c.get("budget", 0)
        name   = c["name"]

        if budget > 0:
            pct    = actual / budget
            filled = min(int(pct * BAR_W), BAR_W)
            bar    = "█" * filled + "░" * (BAR_W - filled)
            bar_label = f"{fmt(actual)} / {fmt(budget)} ₪  ({int(pct * 100)}%)"
            if actual > budget:
                icon = "🔴"
                overruns.append((name, actual - budget, budget))
            elif pct >= 0.85:
                icon = "🟡"
            else:
                icon = "🟢"
        else:
            filled    = max(1, int((actual / max_val) * BAR_W))
            bar       = "█" * filled + "░" * (BAR_W - filled)
            bar_label = f"{fmt(actual)} ₪"
            icon      = "▪️"

        lines.append(f"{RLM}{icon} *{name}*")
        lines.append(f"   {bar} {bar_label}")
        if budget > 0 and actual > budget:
            lines.append(f"{RLM}   ⚠️ חריגה: +{fmt(actual - budget)} ₪")
        elif budget > 0 and pct >= 0.85 and actual <= budget:
            lines.append(f"{RLM}   🟡 נותר: {fmt(int(budget - actual))} ₪")

    lines.append("━━━━━━━━━━━━━━━━━━━━━━")

    if overruns:
        overruns.sort(key=lambda x: -x[1])
        lines.append("\n⚠️ *חריגות מהתקציב:*")
        for name, over, budget in overruns:
            pct_over = (over / budget * 100)
            lines.append(f"  • {name}: +{fmt(over)} ₪  ({pct_over:.0f}% מעל)")

    top3 = [c["name"] for c in active[:3]]
    lines.append(f"\n🎯 *לאן לשים דגש:* {', '.join(top3)}")

    await update.message.reply_text("\n".join(lines))


async def cmd_upload(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Accept an Excel file sent to the bot and save it as the budget file."""
    saved_id = load_chat_id()
    if saved_id and update.effective_chat.id != saved_id:
        return  # ignore unknown senders

    doc = update.message.document if update.message else None
    if not doc or not doc.file_name.endswith(".xlsx"):
        await update.message.reply_text("📎 שלח קובץ Excel (.xlsx) כקובץ מצורף")
        return

    excel_path = os.getenv("EXCEL_PATH", "")
    if not excel_path:
        await update.message.reply_text("⚠️ EXCEL_PATH לא מוגדר")
        return

    tmp_path = excel_path + ".upload.xlsx"  # must end in .xlsx so openpyxl accepts it
    try:
        tg_file = await context.bot.get_file(doc.file_id)
        await tg_file.download_to_drive(tmp_path)

        # Validate before replacing — catch corrupted uploads early
        import openpyxl as _openpyxl
        try:
            _openpyxl.load_workbook(tmp_path, read_only=True).close()
        except Exception as val_err:
            os.unlink(tmp_path)
            await update.message.reply_text(f"⚠️ הקובץ שנשלח לא תקין (xlsx פגום): {val_err}")
            return

        os.replace(tmp_path, excel_path)  # atomic swap — no partial-write corruption
        excel_handler._cat_cache.clear()
        size_kb = doc.file_size / 1024
        await update.message.reply_text(
            f"✅ קובץ האקסל עודכן בהצלחה ({size_kb:.0f} KB)\n"
            f"השתמש ב-/status לבדוק שהנתונים נכונים"
        )
    except Exception as e:
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        await update.message.reply_text(f"⚠️ שגיאה בהעלאת הקובץ: {e}")


async def cmd_export(update: Update, context: ContextTypes.DEFAULT_TYPE):
    excel_path = os.getenv("EXCEL_PATH", "")
    if not excel_path or not os.path.exists(excel_path):
        await update.message.reply_text("⚠️ קובץ האקסל לא נמצא")
        return
    await update.message.reply_document(
        document=open(excel_path, "rb"),
        filename="תקציב_משפחתי_2026.xlsx",
        caption="📊 קובץ תקציב משפחתי עדכני",
    )


async def cmd_status(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        expenses, income, month = excel_handler.get_status()
    except Exception as e:
        await update.message.reply_text(f"⚠️ שגיאה בקריאת הנתונים: {e}")
        return
    if not expenses and not income:
        await update.message.reply_text("לא נמצאו נתונים לחודש הנוכחי")
        return
    await update.message.reply_text(format_status(expenses, income, month), parse_mode="Markdown")


async def cmd_backup(update: Update, context: ContextTypes.DEFAULT_TYPE):
    excel_path = os.getenv("EXCEL_PATH", "")
    if not excel_path or not os.path.exists(excel_path):
        await update.message.reply_text("⚠️ קובץ האקסל לא נמצא")
        return

    # Diagnostic: show which Drive vars are actually visible to the process
    drive_vars = {
        "GOOGLE_CLIENT_ID":     os.getenv("GOOGLE_CLIENT_ID"),
        "GOOGLE_CLIENT_SECRET": os.getenv("GOOGLE_CLIENT_SECRET"),
        "GOOGLE_REFRESH_TOKEN": os.getenv("GOOGLE_REFRESH_TOKEN"),
        "GOOGLE_DRIVE_FOLDER_ID": os.getenv("GOOGLE_DRIVE_FOLDER_ID"),
    }
    missing = [k for k, v in drive_vars.items() if not v]
    logging.info("Drive vars visible to process: %s",
                 {k: ("SET" if v else "MISSING") for k, v in drive_vars.items()})

    if missing:
        await update.message.reply_text(
            f"⚠️ משתני סביבה חסרים לבוט (לא מוגדרים ב-/data/.env):\n"
            + "\n".join(f"  • {k}" for k in missing)
            + "\n\nהוסף אותם ל-/data/.env ב-Railway Shell"
        )
        return

    await update.message.reply_text("⏳ מגבה ל-Google Drive...")
    import drive_backup
    success, msg = drive_backup.backup_to_drive(excel_path)
    await update.message.reply_text(msg)


async def cmd_budget(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Show category list to update monthly budget targets."""
    try:
        cats = excel_handler.get_categories()
    except Exception as e:
        await update.message.reply_text(f"⚠️ שגיאה בקריאת קטגוריות: {e}")
        return

    with_budget    = sum(1 for c in cats if c.get("budget", 0) > 0)
    without_budget = len(cats) - with_budget
    header = (
        f"💰 *עדכון יעד תקציב*\n"
        f"קטגוריות עם יעד: {with_budget}  |  ללא יעד: {without_budget}\n\n"
        f"בחר קטגוריה לעדכון:"
    )
    kb = make_budget_cat_keyboard(cats, page=0)
    await update.message.reply_text(header, reply_markup=kb, parse_mode="Markdown")


async def cmd_debug(update: Update, context: ContextTypes.DEFAULT_TYPE):
    """Diagnose Excel file: list sheet names and check expected sheets exist."""
    import traceback
    import openpyxl

    excel_path = os.getenv("EXCEL_PATH", "")
    data_dir = os.path.dirname(excel_path) or "/data"
    lines = [f"📁 EXCEL_PATH: {excel_path}"]

    # List all files in /data to surface backups
    try:
        data_files = sorted(os.listdir(data_dir))
        lines.append(f"\n📂 קבצים ב-{data_dir}:")
        for f in data_files:
            fp = os.path.join(data_dir, f)
            sz = os.path.getsize(fp) / 1024
            lines.append(f"  {f}  ({sz:.0f} KB)")
    except Exception as dir_err:
        lines.append(f"  (שגיאה בקריאת תיקייה: {dir_err})")

    if not excel_path:
        lines.append("❌ EXCEL_PATH לא מוגדר")
        await update.message.reply_text("\n".join(lines), parse_mode="Markdown")
        return

    if not os.path.exists(excel_path):
        lines.append(f"❌ הקובץ לא קיים: `{excel_path}`")
        await update.message.reply_text("\n".join(lines), parse_mode="Markdown")
        return

    size_kb = os.path.getsize(excel_path) / 1024
    lines.append(f"📦 גודל: {size_kb:.0f} KB")

    try:
        wb = openpyxl.load_workbook(excel_path, read_only=True)
        sheet_names = wb.sheetnames
        wb.close()
        lines.append(f"\n✅ הקובץ נפתח בהצלחה")
        lines.append(f"📋 גיליונות ({len(sheet_names)}):")
        for s in sheet_names:
            lines.append(f"  • `{repr(s)}`")

        expected_index = "I N D E X "
        if expected_index in sheet_names:
            lines.append(f"\n✅ גיליון INDEX נמצא: `{repr(expected_index)}`")
        else:
            lines.append(f"\n❌ גיליון INDEX חסר! מחפש: `{repr(expected_index)}`")
            idx_like = [s for s in sheet_names if "index" in s.lower() or "INDEX" in s]
            if idx_like:
                lines.append(f"   אולי: {idx_like}")

        current = excel_handler.current_sheet()
        if current in sheet_names:
            lines.append(f"✅ גיליון חודש נוכחי נמצא: `{repr(current)}`")
        else:
            lines.append(f"❌ גיליון חודש נוכחי חסר! מחפש: `{repr(current)}`")

    except Exception as e:
        lines.append(f"\n❌ שגיאה בפתיחת הקובץ:")
        lines.append(f"```\n{traceback.format_exc()}\n```")

    await update.message.reply_text("\n".join(lines), parse_mode="Markdown")


async def handle_message(update: Update, context: ContextTypes.DEFAULT_TYPE):
    text = update.message.text.strip()

    # Budget update flow: waiting for amount after user picked a category
    if context.user_data.get("pending_budget"):
        nums = re.findall(r"\d+(?:\.\d+)?", text)
        if nums:
            amount = float(nums[0])
            pb = context.user_data["pending_budget"]
            kb = make_budget_month_keyboard(pb["row"], amount)
            await update.message.reply_text(
                f"💰 *{pb['name']}* — *{fmt(amount)} ₪*\nבחר חודש לעדכון:",
                reply_markup=kb, parse_mode="Markdown"
            )
        else:
            await update.message.reply_text("נא לשלוח סכום (מספר בלבד, לדוגמה: 4000)")
        return

    numbers = re.findall(r"\d+(?:\.\d+)?", text)

    if not numbers:
        await cmd_status(update, context)
        return

    amount = float(numbers[0])
    cat_query = re.sub(r"\d+(?:\.\d+)?", "", text)
    cat_query = re.sub(r"\b(שקל|שקלים|ש\"ח|שח|₪|על|ב|ל)\b", "", cat_query).strip()

    try:
        cats = excel_handler.get_categories()
    except Exception as e:
        await update.message.reply_text(f"⚠️ שגיאה בקריאת קטגוריות: {e}")
        return

    if cat_query:
        # Try full query first, then word-by-word
        matches = excel_handler.find_categories_multi(cat_query, cats)
        if not matches:
            for word in cat_query.split():
                if len(word) > 1:
                    word_matches = excel_handler.find_categories_multi(word, cats)
                    for item in word_matches:
                        if item not in matches:
                            matches.append(item)
            matches.sort(key=lambda x: -x[0])
            matches = matches[:5]

        if len(matches) == 1:
            # Single match — go straight to payment keyboard
            _, cat = matches[0]
            context.user_data["pending_name"] = cat["name"]
            kb = make_pay_keyboard(cat["row"], amount)
            await update.message.reply_text(
                f"*{cat['name']}* — *{fmt(amount)} ₪*\nבחר סוג תשלום:",
                reply_markup=kb, parse_mode="Markdown"
            )
            return
        elif len(matches) > 1:
            # Multiple matches — let user pick
            btns = [
                [InlineKeyboardButton(cat["name"][:30], callback_data=f"cat:{cat['row']}:{amount}")]
                for _, cat in matches
            ]
            btns.append([InlineKeyboardButton("📋 כל הקטגוריות", callback_data=f"page:0:{amount}")])
            await update.message.reply_text(
                f"מצאתי כמה קטגוריות עבור *{fmt(amount)} ₪* — בחר:",
                reply_markup=InlineKeyboardMarkup(btns), parse_mode="Markdown"
            )
            return

    kb = make_cat_keyboard(cats, amount, page=0)
    await update.message.reply_text(
        f"בחר קטגוריה עבור *{fmt(amount)} ₪*:",
        reply_markup=kb, parse_mode="Markdown"
    )


async def _do_add_expense(query, context, month_num, pay_type, row, amount, name, pay_label, sheet_name):
    """Save an expense to Excel, confirm to sender, and forward to other users."""
    sender_id = query.message.chat_id
    await query.edit_message_text(f"⏳ שומר {fmt(amount)} ₪...")

    loop = asyncio.get_event_loop()
    try:
        result = await loop.run_in_executor(None, excel_handler.add_expense, row, amount, sheet_name)
    except Exception as exc:
        logging.error("add_expense failed: %s", exc)
        await query.edit_message_text(f"⚠️ שגיאה בשמירה לאקסל: {exc}")
        return

    await query.edit_message_text(
        format_expense_result(name, amount, result, pay_type, sheet_name),
        parse_mode="Markdown",
    )

    sender_name  = get_user_name(sender_id) or "בן/בת זוג"
    budget       = result.get("budget", 0)
    remaining    = result.get("remaining", 0)
    overrun_note = (
        f"\n⚠️ *חריגה של {fmt(abs(remaining))} ₪ מהתקציב ב-{name}!*"
        if budget > 0 and remaining < 0 else ""
    )
    notif = (
        f"📢 *{sender_name}* הוסיף/ה: *{fmt(amount)} ₪*"
        f" ל-*{name}* ({sheet_name.strip()}) — {pay_label}{overrun_note}"
    )
    for other_id in load_all_chat_ids():
        if other_id != sender_id:
            try:
                await context.bot.send_message(chat_id=other_id, text=notif, parse_mode="Markdown")
            except Exception as e:
                logging.warning("Forward notification to %s failed: %s", other_id, e)


async def handle_callback(update: Update, context: ContextTypes.DEFAULT_TYPE):
    query = update.callback_query
    try:
        await query.answer()
    except Exception:
        return  # callback expired (>10 min old)
    data = query.data

    if data == "cancel":
        context.user_data.pop("pending_name", None)
        await query.edit_message_text("❌ בוטל")
        return

    if data.startswith("cat:"):
        # cat:row:amount → show pay type keyboard
        parts  = data.split(":")
        row    = int(parts[1])
        amount = float(parts[2])
        try:
            cats = excel_handler.get_categories()
        except Exception as e:
            await query.edit_message_text(f"⚠️ {e}")
            return
        cat  = next((c for c in cats if c["row"] == row), None)
        name = cat["name"] if cat else f"שורה {row}"
        context.user_data["pending_name"] = name
        kb = make_pay_keyboard(row, amount)
        await query.edit_message_text(
            f"*{name}* — *{fmt(amount)} ₪*\nבחר סוג תשלום:",
            reply_markup=kb, parse_mode="Markdown"
        )

    elif data.startswith("pay:"):
        # pay:type:row:amount → show month keyboard
        parts     = data.split(":")
        pay_type  = parts[1]
        row       = int(parts[2])
        amount    = float(parts[3])
        name      = context.user_data.get("pending_name", f"שורה {row}")
        pay_label = PAY_LABELS.get(pay_type, pay_type)
        kb = make_month_keyboard(pay_type, row, amount)
        await query.edit_message_text(
            f"*{name}* — *{fmt(amount)} ₪* ({pay_label})\nבחר חודש:",
            reply_markup=kb, parse_mode="Markdown"
        )

    elif data.startswith("month:"):
        # month:month_num:pay_type:row:amount → check budget, then add expense
        parts      = data.split(":")
        month_num  = int(parts[1])
        pay_type   = parts[2]
        row        = int(parts[3])
        amount     = float(parts[4])
        sheet_name = excel_handler.MONTH_SHEETS[month_num]
        name       = context.user_data.pop("pending_name", f"שורה {row}")
        pay_label  = PAY_LABELS.get(pay_type, pay_type)

        # Budget pre-check: warn if this expense pushes close to or over the target
        try:
            cats = excel_handler.get_categories(sheet_name)
            cat  = next((c for c in cats if c["row"] == row), None)
        except Exception:
            cat = None

        if cat and cat.get("budget", 0) > 0:
            budget    = cat["budget"]
            actual    = cat["actual"]
            after     = actual + amount
            pct_after = after / budget

            confirm_btn = InlineKeyboardButton(
                "✅ הוסף בכל זאת" if after > budget else "✅ הוסף",
                callback_data=f"confirmexp:{month_num}:{pay_type}:{row}:{amount}",
            )
            cancel_btn = InlineKeyboardButton("❌ ביטול", callback_data="cancel")

            if after > budget:
                await query.edit_message_text(
                    f"🔴 *שים לב — חריגה מהתקציב!*\n\n"
                    f"קטגוריה: *{name}*  ({sheet_name.strip()})\n"
                    f"💰 יעד:          {fmt(budget)} ₪\n"
                    f"📊 בוצע עד כה:  {fmt(actual)} ₪\n"
                    f"➕ הוספה:        {fmt(amount)} ₪\n"
                    f"─────────────────\n"
                    f"🔴 סה\"כ אחרי: *{fmt(after)} ₪*  (חריגה של *{fmt(after - budget)} ₪*)\n\n"
                    f"להמשיך בכל זאת?",
                    reply_markup=InlineKeyboardMarkup([[confirm_btn], [cancel_btn]]),
                    parse_mode="Markdown",
                )
                return

            if pct_after >= 0.75:
                await query.edit_message_text(
                    f"🟡 *שים לב — מתקרב לגבול התקציב*\n\n"
                    f"קטגוריה: *{name}*  ({sheet_name.strip()})\n"
                    f"💰 יעד:          {fmt(budget)} ₪\n"
                    f"📊 בוצע עד כה:  {fmt(actual)} ₪\n"
                    f"➕ הוספה:        {fmt(amount)} ₪\n"
                    f"─────────────────\n"
                    f"🟡 סה\"כ אחרי: *{fmt(after)} ₪*  (נותר *{fmt(budget - after)} ₪*)\n\n"
                    f"להמשיך?",
                    reply_markup=InlineKeyboardMarkup([[confirm_btn], [cancel_btn]]),
                    parse_mode="Markdown",
                )
                return

        await _do_add_expense(query, context, month_num, pay_type, row, amount, name, pay_label, sheet_name)

    elif data.startswith("confirmexp:"):
        # User confirmed after seeing overrun / near-budget warning
        parts      = data.split(":")
        month_num  = int(parts[1])
        pay_type   = parts[2]
        row        = int(parts[3])
        amount     = float(parts[4])
        sheet_name = excel_handler.MONTH_SHEETS[month_num]
        pay_label  = PAY_LABELS.get(pay_type, pay_type)
        try:
            cats = excel_handler.get_categories(sheet_name)
            cat  = next((c for c in cats if c["row"] == row), None)
            name = cat["name"] if cat else f"שורה {row}"
        except Exception:
            name = f"שורה {row}"
        await _do_add_expense(query, context, month_num, pay_type, row, amount, name, pay_label, sheet_name)

    elif data.startswith("setbud:"):
        row = int(data.split(":")[1])
        try:
            cats = excel_handler.get_categories()
        except Exception as e:
            await query.edit_message_text(f"⚠️ {e}")
            return
        cat = next((c for c in cats if c["row"] == row), None)
        name   = cat["name"] if cat else f"שורה {row}"
        budget = cat.get("budget", 0) if cat else 0
        context.user_data["pending_budget"] = {"row": row, "name": name}
        current_str = f"\nיעד נוכחי: *{fmt(budget)} ₪*" if budget > 0 else "\nעדיין לא הוגדר יעד"
        await query.edit_message_text(
            f"💰 *{name}*{current_str}\n\nשלח את הסכום החדש (₪):",
            parse_mode="Markdown"
        )

    elif data.startswith("budpage:"):
        page = int(data.split(":")[1])
        try:
            cats = excel_handler.get_categories()
        except Exception as e:
            await query.edit_message_text(f"⚠️ {e}")
            return
        kb = make_budget_cat_keyboard(cats, page=page)
        await query.edit_message_reply_markup(kb)

    elif data.startswith("budmonth:"):
        parts      = data.split(":")
        month_num  = int(parts[1])
        row        = int(parts[2])
        amount     = float(parts[3])
        sheet_name = excel_handler.MONTH_SHEETS[month_num]
        pb         = context.user_data.pop("pending_budget", None)
        name       = pb["name"] if pb else f"שורה {row}"

        await query.edit_message_text(f"⏳ מעדכן תקציב...")

        loop = asyncio.get_event_loop()
        try:
            await loop.run_in_executor(None, excel_handler.set_budget, row, amount, sheet_name)
        except Exception as exc:
            logging.error("set_budget failed: %s", exc)
            await query.edit_message_text(f"⚠️ שגיאה בשמירת תקציב: {exc}")
            return

        await query.edit_message_text(
            f"✅ *תקציב עודכן*\n"
            f"קטגוריה: *{name}*\n"
            f"חודש: {sheet_name.strip()}\n"
            f"יעד חדש: *{fmt(amount)} ₪*",
            parse_mode="Markdown"
        )

    elif data.startswith("page:"):
        parts = data.split(":")
        try:
            cats = excel_handler.get_categories()
        except Exception as e:
            await query.edit_message_text(f"⚠️ {e}")
            return
        kb = make_cat_keyboard(cats, float(parts[2]), page=int(parts[1]))
        await query.edit_message_reply_markup(kb)


async def error_handler(update: object, context: ContextTypes.DEFAULT_TYPE):
    logging.error("Exception in handler: %s", context.error, exc_info=context.error)


async def on_startup(app):
    """Sync Desktop→local and warm the category cache before polling starts."""
    desktop = os.getenv("DESKTOP_EXCEL_PATH", "")
    local   = os.getenv("EXCEL_PATH", "")
    if desktop and local and os.path.exists(desktop):
        try:
            shutil.copy2(desktop, local)
            logging.info("סונכרן קובץ Excel מהדסקטופ")
        except Exception as e:
            logging.warning("סנכרון מהדסקטופ נכשל: %s", e)
    loop = asyncio.get_event_loop()
    try:
        await loop.run_in_executor(None, excel_handler.get_categories)
        logging.info("קטגוריות נטענו לזיכרון")
    except Exception as e:
        logging.warning("טעינת קטגוריות נכשלה: %s", e, exc_info=True)


def main():
    if not TOKEN:
        raise RuntimeError("חסר TELEGRAM_BOT_TOKEN ב-.env")
    app = Application.builder().token(TOKEN).post_init(on_startup).build()
    app.add_handler(CommandHandler("start",     cmd_start))
    app.add_handler(CommandHandler("status",    cmd_status))
    app.add_handler(CommandHandler("dashboard", cmd_dashboard))
    app.add_handler(CommandHandler("budget",    cmd_budget))
    app.add_handler(CommandHandler("export",    cmd_export))
    app.add_handler(CommandHandler("backup",    cmd_backup))
    app.add_handler(CommandHandler("debug",     cmd_debug))
    app.add_handler(MessageHandler(filters.Document.FileExtension("xlsx"), cmd_upload))
    app.add_handler(MessageHandler(filters.TEXT & ~filters.COMMAND, handle_message))
    app.add_handler(CallbackQueryHandler(handle_callback))
    app.add_error_handler(error_handler)

    if app.job_queue:
        app.job_queue.run_daily(send_daily_report, time=dtime(8, 0, tzinfo=ISRAEL_TZ))
        print("דוח יומי מוגדר לשעה 08:00 שעון ישראל")
    else:
        print("אזהרה: JobQueue לא זמין - התקן apscheduler")

    print("הבוט פועל...")
    app.run_polling(drop_pending_updates=True)


if __name__ == "__main__":
    main()
