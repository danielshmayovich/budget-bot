import os
import re
import time
import shutil
import logging
from datetime import datetime
from difflib import SequenceMatcher
import openpyxl
from dotenv import load_dotenv

load_dotenv()

EXCEL_PATH   = os.getenv("EXCEL_PATH", "")
DESKTOP_PATH = os.getenv("DESKTOP_EXCEL_PATH", "")

MONTH_SHEETS = {
    1: "ינואר  ", 2: "פברואר  ", 3: "מרץ ", 4: "אפריל  ",
    5: "מאי ", 6: "יוני ", 7: "יולי ", 8: "אוגוסט ",
    9: "ספטמבר ", 10: "אוקטובר ", 11: "נובמבר  ", 12: "דצמבר "
}

EXPENSE_START_ROW = 6
EXPENSE_END_ROW   = 111
INCOME_START_ROW  = 117
CAT_END_ROW       = 200
COL_NAME          = 3
COL_BUDGET        = 4  # column D
COL_EXP_START     = 5
COL_EXP_END       = 35
IDX_ROW_RE        = re.compile(r"!B(\d+)")

# Cache categories per sheet to avoid repeated slow reads
_cat_cache: dict = {}  # sheet_name -> (cats, timestamp)
CACHE_TTL = 120  # seconds


def current_sheet():
    return MONTH_SHEETS[datetime.now().month]


def _parse_name(cell_val, idx_ws):
    if isinstance(cell_val, str) and cell_val.startswith("="):
        m = IDX_ROW_RE.search(cell_val)
        if m:
            return idx_ws.cell(int(m.group(1)), 2).value
        return None
    return cell_val


def _sum_entries(ws, row):
    return sum(
        float(ws.cell(row, c).value)
        for c in range(COL_EXP_START, COL_EXP_END + 1)
        if isinstance(ws.cell(row, c).value, (int, float))
    )


def _load_wb(write=False):
    try:
        return openpyxl.load_workbook(EXCEL_PATH)
    except Exception as e:
        raise IOError(
            f"לא ניתן לפתוח את קובץ ה-Excel - ייתכן שנשמר כרגע, נסה שוב ({type(e).__name__})"
        ) from e


def _build_row_list(ws, idx, start_row, end_row):
    items = []
    for r in range(start_row, end_row + 1):
        name = _parse_name(ws.cell(r, COL_NAME).value, idx)
        if not name or str(name).strip() in ("", "#REF!", 'סה"כ כללי', "הכנסות", "סעיפי הוצאה "):
            continue
        total = _sum_entries(ws, r)
        budget_val = ws.cell(r, COL_BUDGET).value
        budget = float(budget_val) if isinstance(budget_val, (int, float)) and budget_val > 0 else 0
        items.append({"row": r, "name": str(name).strip(), "actual": total, "budget": budget})
    return items


def get_categories(sheet_name=None):
    if not sheet_name:
        sheet_name = current_sheet()
    now = time.time()
    cached = _cat_cache.get(sheet_name)
    if cached and now - cached[1] < CACHE_TTL:
        return cached[0]

    wb = _load_wb(write=False)
    try:
        ws  = wb[sheet_name]
        idx = wb["I N D E X "]
        cats = _build_row_list(ws, idx, EXPENSE_START_ROW, EXPENSE_END_ROW)
    finally:
        wb.close()

    for c in cats:
        c.setdefault("remaining", 0)
    _cat_cache[sheet_name] = (cats, now)
    return cats


def get_status(sheet_name=None):
    if not sheet_name:
        sheet_name = current_sheet()
    wb = _load_wb(write=False)
    try:
        ws  = wb[sheet_name]
        idx = wb["I N D E X "]
        expenses = _build_row_list(ws, idx, EXPENSE_START_ROW, EXPENSE_END_ROW)
        income   = _build_row_list(ws, idx, INCOME_START_ROW, CAT_END_ROW)
    finally:
        wb.close()

    active_expenses = [c for c in expenses if c["actual"] > 0]
    active_income   = [c for c in income   if c["actual"] > 0]
    return active_expenses, active_income, sheet_name.strip()


def find_category(query, cats=None):
    if cats is None:
        cats = get_categories()
    q = query.strip()
    best, best_score = None, 0
    for cat in cats:
        name = cat["name"]
        if q in name or name.startswith(q):
            score = 0.85
        else:
            score = SequenceMatcher(None, q, name).ratio()
        if score > best_score:
            best_score = score
            best = cat
    return best, best_score


def find_categories_multi(query, cats=None, threshold=0.5, max_results=5):
    """Return all categories that match query above threshold, sorted by score desc."""
    if cats is None:
        cats = get_categories()
    q = query.strip()
    scored = []
    for cat in cats:
        name = cat["name"]
        if q in name or name.startswith(q):
            score = 0.85
        else:
            score = SequenceMatcher(None, q, name).ratio()
        if score >= threshold:
            scored.append((score, cat))
    scored.sort(key=lambda x: -x[0])
    return [(score, cat) for score, cat in scored[:max_results]]


def _wb_save_safe(wb, path):
    """Save workbook, stripping charts that openpyxl can't round-trip.

    Legacy charts crash openpyxl serialization: 'str' has no attribute 'to_tree'.
    wb.worksheets misses ChartSheet objects; wb._sheets covers everything.
    """
    for ws in wb._sheets:
        n = len(ws._charts) if hasattr(ws, "_charts") else 0
        if n:
            logging.info("Stripping %d chart(s) from sheet %r", n, ws.title)
        if hasattr(ws, "_charts"):
            ws._charts = []
        if hasattr(ws, "_images"):
            ws._images = []

    try:
        wb.save(path)
        return
    except AttributeError as e:
        if "to_tree" not in str(e):
            raise
        logging.warning("to_tree after chart strip — full traceback:", exc_info=True)

    from openpyxl.workbook.defined_name import DefinedNameList
    wb.defined_names = DefinedNameList()
    logging.warning("Also stripped defined_names — retrying save")
    wb.save(path)


def add_expense(row_idx, amount, sheet_name=None):
    if not sheet_name:
        sheet_name = current_sheet()
    wb = _load_wb(write=True)
    ws = wb[sheet_name]
    next_col = None
    for c in range(COL_EXP_START, COL_EXP_END + 1):
        if not isinstance(ws.cell(row_idx, c).value, (int, float)):
            next_col = c
            break
    if next_col is None:
        wb.close()
        raise ValueError("אין מקום פנוי בשורת הקטגוריה")

    ws.cell(row_idx, next_col, float(amount))
    actual = _sum_entries(ws, row_idx)
    tmp_path = EXCEL_PATH + ".tmp"
    try:
        _wb_save_safe(wb, tmp_path)
        os.replace(tmp_path, EXCEL_PATH)  # atomic: no partial-write corruption
    except PermissionError:
        wb.close()
        raise PermissionError("הקובץ נעול - סגור את Excel ונסה שוב")
    except Exception as e:
        logging.error("שגיאה בשמירת קובץ Excel", exc_info=True)
        wb.close()
        try:
            os.unlink(tmp_path)
        except OSError:
            pass
        raise IOError(f"שגיאה בשמירת הקובץ: {type(e).__name__} - {e}") from e
    wb.close()

    # Invalidate cache so next read reflects the new entry
    _cat_cache.pop(sheet_name, None)

    # Copy updated local file back to Desktop (slow, but caller runs this in background)
    if DESKTOP_PATH:
        try:
            shutil.copy2(EXCEL_PATH, DESKTOP_PATH)
        except Exception as e:
            logging.warning("Desktop sync failed: %s", e)

    return {"budget": 0, "actual": actual, "remaining": 0}
